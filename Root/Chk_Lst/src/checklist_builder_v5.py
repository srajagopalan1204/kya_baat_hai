#!/usr/bin/env python3
"""
checklist_builder_v5.py

Injects Header + Steps from an XLSX spec into a v8-style HTML template.

Template requirements:
  - contains: let sopInfo = { ... };
  - contains: let steps = [ ... ];

Header sheet supported layouts:
  A) Key/Value pairs table (columns named Key/Value OR uses A/B)
  B) Field names across a row, values on the next row

Steps sheet supported:
  - Finds the header row by scanning for known column names
  - Reads rows under that header until blank streak

This version avoids JS syntax issues by injecting JSON (newlines become \\n).
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


# -----------------------------
# Utilities
# -----------------------------

def norm(v: Any) -> str:
    return "" if v is None else str(v).strip()

def low(v: Any) -> str:
    return norm(v).lower()

def boolish(v: Any, default: bool = False) -> bool:
    s = low(v)
    if s in ("y", "yes", "true", "1", "done", "complete", "completed"):
        return True
    if s in ("n", "no", "false", "0", ""):
        return False
    return default

def json_for_js(obj: Any) -> str:
    # Safe JS injection: embed as JSON literal directly in script
    return json.dumps(obj, ensure_ascii=False, indent=2)

def find_sheet_name(wb, candidates: List[str]) -> Optional[str]:
    existing = {name.lower(): name for name in wb.sheetnames}
    for c in candidates:
        if c.lower() in existing:
            return existing[c.lower()]
    return None


# -----------------------------
# Header parsing
# -----------------------------

HEADER_KEY_SYNONYMS = {
    "name": {"name", "sop name", "sopname", "sop_nm"},
    "id": {"id", "sop id", "sopid", "sop_id"},
    "entity": {"entity", "sop entity", "sopentity", "sop_entity"},
    "repo": {"repo", "repo path", "metarepo", "meta repo", "codespaces repo"},
    "webRoot": {"webroot", "web root", "publish", "publish target", "stage repo", "web repo"},
    "runLabel": {"runlabel", "run label"},
    "imgFolder": {"imgfolder", "img folder", "sopimgfolder", "image folder"},
    "templateTag": {"templatetag", "template tag", "tag"},
}

def canonical_header_key(k: str) -> Optional[str]:
    lk = low(k)
    for canon, syns in HEADER_KEY_SYNONYMS.items():
        if lk in syns:
            return canon
    return None

def read_header_key_value(ws) -> Dict[str, str]:
    """
    Reads Header sheet as key/value pairs.
    Looks for columns named Key/Value in row 1; otherwise uses A/B.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    r1 = [norm(x) for x in rows[0]]
    idx_key = 0
    idx_val = 1

    # detect Key/Value header columns
    for i, h in enumerate(r1):
        if low(h) in ("key", "field", "header_key", "name"):
            idx_key = i
        if low(h) in ("value", "val", "header_value"):
            idx_val = i

    out: Dict[str, str] = {}
    blank_streak = 0

    for r in rows[1:]:
        key = norm(r[idx_key]) if idx_key < len(r) else ""
        val = norm(r[idx_val]) if idx_val < len(r) else ""

        if key == "":
            blank_streak += 1
            if blank_streak >= 5:
                break
            continue

        blank_streak = 0
        out[key] = val

    return out

def read_header_row_values(ws) -> Dict[str, str]:
    """
    Reads Header sheet where field names are in one row and values in the next row.
    We scan first ~15 rows to find a row with multiple known header fields.
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {}

    best_i = None
    best_hits = 0

    # scan top rows to find the "field names" row
    for i in range(min(15, len(rows) - 1)):
        fields = [norm(x) for x in rows[i]]
        hits = 0
        for f in fields:
            if canonical_header_key(f):
                hits += 1
        if hits > best_hits:
            best_hits = hits
            best_i = i

    if best_i is None or best_hits < 3:
        return {}

    field_row = [norm(x) for x in rows[best_i]]
    value_row = [norm(x) for x in rows[best_i + 1]]

    out: Dict[str, str] = {}
    for j, f in enumerate(field_row):
        canon = canonical_header_key(f)
        if not canon:
            continue
        if j < len(value_row) and norm(value_row[j]) != "":
            out[canon] = norm(value_row[j])

    return out

def build_sopinfo(header_data: Dict[str, str]) -> Dict[str, Any]:
    """
    Build sopInfo object expected by v8 template.
    """
    base = {
        "name": "",
        "id": "",
        "entity": "",
        "repo": "/workspaces/SOP_Build",
        "webRoot": "/SOP_Stage",
        "runLabel": "",
        "imgFolder": "../outputs/images/<SOP_ID>",
        "templateTag": "v8 – injected"
    }

    # header_data may be mixed (canonical keys OR original keys)
    # normalize: allow either canonical keys or raw keys that match synonyms
    normalized: Dict[str, str] = {}

    for k, v in header_data.items():
        ck = canonical_header_key(k)
        if ck:
            normalized[ck] = norm(v)
        else:
            # already canonical?
            if k in base:
                normalized[k] = norm(v)

    for k in base:
        if k in normalized and normalized[k] != "":
            base[k] = normalized[k]

    return base


# -----------------------------
# Steps parsing
# -----------------------------

STEP_COL_SYNONYMS = {
    "order": {"order", "step", "step_no", "step number", "seq", "sequence"},
    "id": {"id", "step_id", "code"},
    "title": {"title", "step_title", "name"},
    "command": {"command", "cmd", "procedure", "instructions"},
    "reminder": {"reminder", "hints", "hint", "tips", "tip"},
    "notes": {"notes", "run_notes", "comments"},
    "done": {"done", "status", "complete", "completed"},
}

def canon_step_col(name: str) -> Optional[str]:
    ln = low(name)
    for canon, syns in STEP_COL_SYNONYMS.items():
        if ln in syns:
            return canon
    return None

def find_steps_header_row(rows: List[Tuple[Any, ...]]) -> Optional[int]:
    """
    Scan first ~30 rows for a header that contains at least 3 known step columns.
    """
    best_i = None
    best_hits = 0
    for i in range(min(30, len(rows))):
        r = [norm(x) for x in rows[i]]
        hits = 0
        for cell in r:
            if canon_step_col(cell):
                hits += 1
        if hits > best_hits:
            best_hits = hits
            best_i = i
    if best_i is None or best_hits < 3:
        return None
    return best_i

def read_steps(ws) -> List[Dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    hdr_i = find_steps_header_row(rows)
    if hdr_i is None:
        return []

    header = [norm(x) for x in rows[hdr_i]]
    col_map: Dict[str, int] = {}
    for j, h in enumerate(header):
        c = canon_step_col(h)
        if c and c not in col_map:
            col_map[c] = j

    out: List[Dict[str, Any]] = []
    blank_streak = 0

    for r in rows[hdr_i + 1:]:
        if all(norm(x) == "" for x in r):
            blank_streak += 1
            if blank_streak >= 10:
                break
            continue
        blank_streak = 0

        def get(col: str) -> str:
            if col not in col_map:
                return ""
            j = col_map[col]
            return norm(r[j]) if j < len(r) else ""

        order_raw = get("order")
        try:
            order = int(order_raw) if order_raw != "" else len(out) + 1
        except Exception:
            order = len(out) + 1

        step = {
            "id": get("id") or f"Step{order}",
            "order": order,
            "title": get("title") or f"Step {order}",
            "command": get("command"),
            "reminder": get("reminder"),
            "notes": get("notes"),
            "done": boolish(get("done"), default=False),
            "runs": []
        }
        out.append(step)

    out.sort(key=lambda x: int(x.get("order", 10**9)))
    return out


# -----------------------------
# Template injection
# -----------------------------

SOPINFO_RE = re.compile(r"(?P<prefix>\blet\s+sopInfo\s*=\s*)\{.*?\}\s*;", re.DOTALL)
STEPS_RE   = re.compile(r"(?P<prefix>\blet\s+steps\s*=\s*)\[\s*.*?\s*\]\s*;", re.DOTALL)

def inject(template_html: str, sop_info: Dict[str, Any], steps: List[Dict[str, Any]]) -> str:
    if not SOPINFO_RE.search(template_html):
        raise RuntimeError("Template missing: let sopInfo = { ... };")
    if not STEPS_RE.search(template_html):
        raise RuntimeError("Template missing: let steps = [ ... ];")

    sop_js = json_for_js(sop_info)
    steps_js = json_for_js(steps)

    template_html = SOPINFO_RE.sub(lambda m: f"{m.group('prefix')}{sop_js};", template_html, count=1)
    template_html = STEPS_RE.sub(lambda m: f"{m.group('prefix')}{steps_js};", template_html, count=1)

    return template_html


# -----------------------------
# Main
# -----------------------------

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--spec", required=True)
    ap.add_argument("--template", required=True)
    ap.add_argument("--out-html", required=True)
    ap.add_argument("--debug", action="store_true")
    args = ap.parse_args()

    spec_path = Path(args.spec)
    tpl_path = Path(args.template)
    out_path = Path(args.out_html)

    if not spec_path.exists():
        print(f"ERROR: spec not found: {spec_path}", file=sys.stderr)
        return 2
    if not tpl_path.exists():
        print(f"ERROR: template not found: {tpl_path}", file=sys.stderr)
        return 2

    wb = load_workbook(spec_path, data_only=True)

    sh_header = find_sheet_name(wb, ["Header", "META", "Meta"])
    sh_steps  = find_sheet_name(wb, ["Steps", "Checklist", "STEPS"])

    if not sh_header:
        print("ERROR: missing sheet 'Header' (or 'Meta').", file=sys.stderr)
        return 2
    if not sh_steps:
        print("ERROR: missing sheet 'Steps' (or 'Checklist').", file=sys.stderr)
        return 2

    ws_h = wb[sh_header]
    ws_s = wb[sh_steps]

    # Try both header layouts and merge results (row-values wins if it finds keys)
    kv = read_header_key_value(ws_h)
    rv = read_header_row_values(ws_h)

    # Merge: if rv has entries, prefer them
    header_data: Dict[str, str] = {}
    header_data.update(kv)       # raw kv
    header_data.update(rv)       # canonical row-values

    sop_info = build_sopinfo(header_data)
    steps = read_steps(ws_s)

    template_html = tpl_path.read_text(encoding="utf-8", errors="replace")
    out_html = inject(template_html, sop_info, steps)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(out_html, encoding="utf-8")

    if args.debug:
        # show what we actually got
        print(f"[debug] Header sheet: {sh_header}")
        print(f"[debug] Steps sheet : {sh_steps}")
        print(f"[debug] header kv pairs read: {len(kv)}")
        print(f"[debug] header row-values read: {len(rv)}")
        print(f"[debug] sopInfo.id='{sop_info.get('id','')}' sopInfo.name='{sop_info.get('name','')}'")
        print(f"[debug] steps read: {len(steps)}")
        if steps:
            print(f"[debug] first step id='{steps[0].get('id')}' title='{steps[0].get('title')}'")

    print(f"OK: wrote {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
